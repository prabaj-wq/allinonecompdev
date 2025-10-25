from fastapi import APIRouter, HTTPException, status, Query, Request
from fastapi.responses import FileResponse, StreamingResponse
from typing import List, Optional, Dict, Any
from pydantic import BaseModel
import psycopg2
import os
import json
import pandas as pd
from datetime import datetime
from pathlib import Path
import tempfile
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(prefix="/financial-statements", tags=["Financial Statements"])

class GenerateRequest(BaseModel):
    company_name: str
    process_id: Optional[int] = None
    scenario_id: Optional[int] = None
    hierarchy_id: int
    entity_ids: List[Any]
    period_ids: List[int]
    report_type: str
    show_zero_balances: bool = False
    show_ic_column: bool = True
    show_other_column: bool = True
    rounding_factor: int = 1
    currency: str = "INR"

def get_db_config():
    """Get database configuration"""
    if os.getenv('DOCKER_ENV') == 'true':
        POSTGRES_HOST = os.getenv('POSTGRES_HOST', 'postgres')
    else:
        POSTGRES_HOST = os.getenv('POSTGRES_HOST', 'localhost')
        
    return {
        'host': POSTGRES_HOST,
        'port': os.getenv('POSTGRES_PORT', '5432'),
        'user': 'postgres',
        'password': 'root@123'
    }

@router.post("/generate")
async def generate_financial_statements(req: GenerateRequest):
    """Generate comprehensive financial statements based on account hierarchy"""
    try:
        db_config = get_db_config()
        company_db_name = req.company_name.lower().replace(' ', '_').replace('-', '_')
        
        conn = psycopg2.connect(database=company_db_name, **db_config)
        cur = conn.cursor()
        
        # 1. Fetch hierarchy structure
        cur.execute("""
            SELECT id, code, name, parent_id, level, hierarchy_id
            FROM account_hierarchy_nodes
            WHERE hierarchy_id = %s
            ORDER BY level, code
        """, (req.hierarchy_id,))
        nodes = cur.fetchall()
        
        # 2. Fetch accounts mapped to hierarchy
        cur.execute("""
            SELECT a.account_code, a.account_name, a.node_id
            FROM axes_accounts a
            WHERE a.company_name = %s
        """, (req.company_name,))
        accounts = cur.fetchall()
        
        # 3. Fetch amounts from data input tables
        amounts_data = {}
        
        # Get process table name
        if req.process_id:
            cur.execute("SELECT table_name FROM processes WHERE id = %s", (req.process_id,))
            process_result = cur.fetchone()
            table_prefix = process_result[0] if process_result else "actuals"
        else:
            table_prefix = "actuals"
        
        # Fetch entity amounts
        entity_table = f"{table_prefix}_entity_amounts_entries"
        try:
            period_filter = ",".join(str(p) for p in req.period_ids)
            entity_filter = ",".join(f"'{e}'" for e in req.entity_ids)
            
            query = f"""
                SELECT account_code, entity_id, SUM(CAST(amount AS NUMERIC)) as total_amount
                FROM {entity_table}
                WHERE scenario_id = %s
                AND period_id IN ({period_filter})
                AND entity_id IN ({entity_filter})
                GROUP BY account_code, entity_id
            """
            cur.execute(query, (req.scenario_id,))
            entity_amounts = cur.fetchall()
            
            for acc_code, entity_id, amount in entity_amounts:
                if acc_code not in amounts_data:
                    amounts_data[acc_code] = {}
                if entity_id not in amounts_data[acc_code]:
                    amounts_data[acc_code][entity_id] = {"entity_amount": 0, "ic_amount": 0, "other_amount": 0}
                amounts_data[acc_code][entity_id]["entity_amount"] = float(amount or 0)
        except Exception as e:
            print(f"Error fetching entity amounts: {e}")
        
        # Fetch IC amounts
        if req.show_ic_column:
            ic_table = f"{table_prefix}_ic_amounts_entries"
            try:
                query = f"""
                    SELECT from_account_code as account_code, from_entity_id as entity_id, 
                           SUM(CAST(amount AS NUMERIC)) as total_amount
                    FROM {ic_table}
                    WHERE scenario_id = %s
                    AND period_id IN ({period_filter})
                    AND from_entity_id IN ({entity_filter})
                    GROUP BY from_account_code, from_entity_id
                """
                cur.execute(query, (req.scenario_id,))
                ic_amounts = cur.fetchall()
                
                for acc_code, entity_id, amount in ic_amounts:
                    if acc_code not in amounts_data:
                        amounts_data[acc_code] = {}
                    if entity_id not in amounts_data[acc_code]:
                        amounts_data[acc_code][entity_id] = {"entity_amount": 0, "ic_amount": 0, "other_amount": 0}
                    amounts_data[acc_code][entity_id]["ic_amount"] = float(amount or 0)
            except Exception as e:
                print(f"Error fetching IC amounts: {e}")
        
        # Fetch other amounts
        if req.show_other_column:
            other_table = f"{table_prefix}_other_amounts_entries"
            try:
                query = f"""
                    SELECT account_code, entity_id, SUM(CAST(amount AS NUMERIC)) as total_amount
                    FROM {other_table}
                    WHERE scenario_id = %s
                    AND period_id IN ({period_filter})
                    AND entity_id IN ({entity_filter})
                    GROUP BY account_code, entity_id
                """
                cur.execute(query, (req.scenario_id,))
                other_amounts = cur.fetchall()
                
                for acc_code, entity_id, amount in other_amounts:
                    if acc_code not in amounts_data:
                        amounts_data[acc_code] = {}
                    if entity_id not in amounts_data[acc_code]:
                        amounts_data[acc_code][entity_id] = {"entity_amount": 0, "ic_amount": 0, "other_amount": 0}
                    amounts_data[acc_code][entity_id]["other_amount"] = float(amount or 0)
            except Exception as e:
                print(f"Error fetching other amounts: {e}")
        
        cur.close()
        conn.close()
        
        # 4. Build hierarchy structure with amounts
        nodes_dict = {}
        for node_id, code, name, parent_id, level, hierarchy_id in nodes:
            nodes_dict[node_id] = {
                "id": node_id,
                "code": code,
                "name": name,
                "parent_id": parent_id,
                "level": level,
                "accounts": [],
                "children": []
            }
        
        # Map accounts to nodes
        for acc_code, acc_name, node_id in accounts:
            if node_id and node_id in nodes_dict:
                nodes_dict[node_id]["accounts"].append({
                    "code": acc_code,
                    "name": acc_name,
                    "amounts": amounts_data.get(acc_code, {})
                })
        
        # Build tree structure
        root_nodes = []
        for node_id, node_data in nodes_dict.items():
            if node_data["parent_id"] is None:
                root_nodes.append(node_data)
            else:
                parent_id = node_data["parent_id"]
                if parent_id in nodes_dict:
                    nodes_dict[parent_id]["children"].append(node_data)
        
        return {
            "report_title": f"{req.report_type.replace('_', ' ').title()} - Financial Statement",
            "report_id": f"report_{datetime.now().timestamp()}",
            "currency": req.currency,
            "nodes": root_nodes,
            "entities": req.entity_ids,
            "generated_at": datetime.utcnow().isoformat()
        }
        
    except Exception as e:
        print(f"Error generating financial statements: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate financial statements: {str(e)}"
        )

@router.get("/drill-down")
async def get_drill_down_data(
    company_name: str = Query(...),
    process_id: int = Query(...),
    scenario_id: int = Query(...),
    account_code: str = Query(...),
    entity_id: str = Query(...),
    period_ids: str = Query(...)
):
    """Get drill-down transaction details for an account"""
    try:
        db_config = get_db_config()
        company_db_name = company_name.lower().replace(' ', '_').replace('-', '_')
        
        conn = psycopg2.connect(database=company_db_name, **db_config)
        cur = conn.cursor()
        
        # Get process table name
        cur.execute("SELECT table_name FROM processes WHERE id = %s", (process_id,))
        process_result = cur.fetchone()
        table_prefix = process_result[0] if process_result else "actuals"
        
        period_list = period_ids.split(",")
        period_filter = ",".join(period_list)
        
        # Fetch entries from all three tables
        entries = []
        
        # Entity amounts
        entity_table = f"{table_prefix}_entity_amounts_entries"
        try:
            query = f"""
                SELECT transaction_date, reference_id, description, amount, 'Entity' as type
                FROM {entity_table}
                WHERE account_code = %s
                AND entity_id = %s
                AND scenario_id = %s
                AND period_id IN ({period_filter})
                ORDER BY transaction_date DESC
            """
            cur.execute(query, (account_code, entity_id, scenario_id))
            entries.extend(cur.fetchall())
        except Exception as e:
            print(f"Error fetching entity entries: {e}")
        
        # IC amounts
        ic_table = f"{table_prefix}_ic_amounts_entries"
        try:
            query = f"""
                SELECT transaction_date, reference_id, description, amount, 'IC' as type
                FROM {ic_table}
                WHERE from_account_code = %s
                AND from_entity_id = %s
                AND scenario_id = %s
                AND period_id IN ({period_filter})
                ORDER BY transaction_date DESC
            """
            cur.execute(query, (account_code, entity_id, scenario_id))
            entries.extend(cur.fetchall())
        except Exception as e:
            print(f"Error fetching IC entries: {e}")
        
        # Other amounts
        other_table = f"{table_prefix}_other_amounts_entries"
        try:
            query = f"""
                SELECT transaction_date, reference_id, description, amount, 'Other' as type
                FROM {other_table}
                WHERE account_code = %s
                AND entity_id = %s
                AND scenario_id = %s
                AND period_id IN ({period_filter})
                ORDER BY transaction_date DESC
            """
            cur.execute(query, (account_code, entity_id, scenario_id))
            entries.extend(cur.fetchall())
        except Exception as e:
            print(f"Error fetching other entries: {e}")
        
        cur.close()
        conn.close()
        
        # Format entries
        formatted_entries = []
        for date, ref, desc, amount, entry_type in entries:
            formatted_entries.append({
                "date": str(date) if date else "",
                "reference": ref or "",
                "description": desc or "",
                "amount": float(amount) if amount else 0,
                "type": entry_type
            })
        
        return {
            "account_code": account_code,
            "entity_id": entity_id,
            "entries": formatted_entries
        }
        
    except Exception as e:
        print(f"Error getting drill-down data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get drill-down data: {str(e)}"
        )

@router.post("/export")
async def export_financial_statement(request: Request):
    """Export financial statement to Excel or PDF"""
    try:
        body = await request.body()
        data = json.loads(body)
        
        company_name = data.get("company_name")
        report_data = data.get("report_data")
        config = data.get("config")
        format_type = data.get("format", "excel")
        
        if format_type == "excel":
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Statement"
            
            # Add header
            ws["A1"] = report_data.get("report_title", "Financial Statement")
            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"] = f"Company: {company_name}"
            ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            # Add column headers
            row = 5
            ws[f"A{row}"] = "Account Code"
            ws[f"B{row}"] = "Description"
            col = 3
            
            entities = report_data.get("entities", [])
            for entity in entities:
                ws.cell(row, col).value = str(entity)
                col += 1
            
            # Style headers
            for cell in ws[row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data rows (simplified - would need full hierarchy traversal)
            row += 1
            nodes = report_data.get("nodes", [])
            for node in nodes:
                ws.cell(row, 1).value = node.get("code")
                ws.cell(row, 2).value = node.get("name")
                row += 1
            
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return StreamingResponse(
                excel_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=financial_statement_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        
        else:
            raise HTTPException(status_code=400, detail="Only Excel export is currently supported")
        
    except Exception as e:
        print(f"Error exporting financial statement: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export financial statement: {str(e)}"
        )

def generate_balance_sheet(entries):
    """Generate balance sheet from trial balance entries"""
    assets = []
    liabilities = []
    equity = []
    
    for entry in entries:
        account_code, account_name, debit, credit, balance, account_type, entity_code = entry
        
        entry_dict = {
            'account_code': account_code,
            'account_name': account_name,
            'balance_amount': float(balance) if balance else 0,
            'entity_code': entity_code
        }
        
        if account_type == 'Asset':
            assets.append(entry_dict)
        elif account_type == 'Liability':
            liabilities.append(entry_dict)
        elif account_type == 'Equity':
            equity.append(entry_dict)
    
    # Calculate totals
    total_assets = sum(item['balance_amount'] for item in assets)
    total_liabilities = sum(item['balance_amount'] for item in liabilities)
    total_equity = sum(item['balance_amount'] for item in equity)
    
    return {
        "assets": {
            "current_assets": [a for a in assets if a['account_code'].startswith(('1', '11'))],
            "non_current_assets": [a for a in assets if a['account_code'].startswith(('12', '13', '14', '15'))],
            "total": total_assets
        },
        "liabilities": {
            "current_liabilities": [l for l in liabilities if l['account_code'].startswith(('2', '20', '21'))],
            "non_current_liabilities": [l for l in liabilities if l['account_code'].startswith(('22', '23', '24'))],
            "total": total_liabilities
        },
        "equity": {
            "items": equity,
            "total": total_equity
        },
        "total_liabilities_and_equity": total_liabilities + total_equity
    }

def generate_income_statement(entries):
    """Generate income statement from trial balance entries"""
    revenue = []
    expenses = []
    
    for entry in entries:
        account_code, account_name, debit, credit, balance, account_type, entity_code = entry
        
        entry_dict = {
            'account_code': account_code,
            'account_name': account_name,
            'balance_amount': float(balance) if balance else 0,
            'entity_code': entity_code
        }
        
        if account_type in ['Revenue', 'Income']:
            revenue.append(entry_dict)
        elif account_type == 'Expense':
            expenses.append(entry_dict)
    
    # Calculate totals
    total_revenue = sum(item['balance_amount'] for item in revenue)
    total_expenses = sum(item['balance_amount'] for item in expenses)
    net_income = total_revenue - total_expenses
    
    return {
        "revenue": {
            "items": revenue,
            "total": total_revenue
        },
        "expenses": {
            "cost_of_sales": [e for e in expenses if e['account_code'].startswith(('5', '50'))],
            "operating_expenses": [e for e in expenses if e['account_code'].startswith(('6', '60', '61'))],
            "other_expenses": [e for e in expenses if e['account_code'].startswith(('7', '70'))],
            "total": total_expenses
        },
        "net_income": net_income,
        "gross_profit": total_revenue - sum(e['balance_amount'] for e in expenses if e['account_code'].startswith(('5', '50')))
    }

def generate_cash_flow_statement(entries):
    """Generate simplified cash flow statement"""
    # This is a simplified version - in practice, this would require more detailed analysis
    cash_accounts = [e for e in entries if e[0].startswith(('1000', '1001', '1002'))]  # Cash accounts
    
    total_cash = sum(float(e[4]) if e[4] else 0 for e in cash_accounts)
    
    return {
        "operating_activities": {
            "net_income": 0,  # Would be calculated from income statement
            "adjustments": [],
            "net_cash_from_operations": 0
        },
        "investing_activities": {
            "items": [],
            "net_cash_from_investing": 0
        },
        "financing_activities": {
            "items": [],
            "net_cash_from_financing": 0
        },
        "net_change_in_cash": 0,
        "cash_beginning": 0,
        "cash_ending": total_cash
    }

def generate_sample_statements(period, year, company_name):
    """Generate sample financial statements"""
    return {
        "success": True,
        "period": period,
        "year": year,
        "company_name": company_name,
        "balance_sheet": {
            "assets": {
                "current_assets": [
                    {"account_code": "1000", "account_name": "Cash and Cash Equivalents", "balance_amount": 150000},
                    {"account_code": "1200", "account_name": "Accounts Receivable", "balance_amount": 85000},
                    {"account_code": "1300", "account_name": "Inventory", "balance_amount": 120000}
                ],
                "non_current_assets": [
                    {"account_code": "1500", "account_name": "Property, Plant & Equipment", "balance_amount": 500000},
                    {"account_code": "1600", "account_name": "Intangible Assets", "balance_amount": 75000}
                ],
                "total": 930000
            },
            "liabilities": {
                "current_liabilities": [
                    {"account_code": "2000", "account_name": "Accounts Payable", "balance_amount": 65000},
                    {"account_code": "2100", "account_name": "Short-term Debt", "balance_amount": 50000}
                ],
                "non_current_liabilities": [
                    {"account_code": "2200", "account_name": "Long-term Debt", "balance_amount": 200000}
                ],
                "total": 315000
            },
            "equity": {
                "items": [
                    {"account_code": "3000", "account_name": "Share Capital", "balance_amount": 400000},
                    {"account_code": "3100", "account_name": "Retained Earnings", "balance_amount": 215000}
                ],
                "total": 615000
            },
            "total_liabilities_and_equity": 930000
        },
        "income_statement": {
            "revenue": {
                "items": [
                    {"account_code": "4000", "account_name": "Sales Revenue", "balance_amount": 1200000},
                    {"account_code": "4100", "account_name": "Service Revenue", "balance_amount": 300000}
                ],
                "total": 1500000
            },
            "expenses": {
                "cost_of_sales": [
                    {"account_code": "5000", "account_name": "Cost of Goods Sold", "balance_amount": 720000}
                ],
                "operating_expenses": [
                    {"account_code": "6000", "account_name": "Salaries and Wages", "balance_amount": 350000},
                    {"account_code": "6100", "account_name": "Rent Expense", "balance_amount": 120000},
                    {"account_code": "6200", "account_name": "Utilities", "balance_amount": 45000}
                ],
                "other_expenses": [
                    {"account_code": "7000", "account_name": "Interest Expense", "balance_amount": 15000}
                ],
                "total": 1250000
            },
            "gross_profit": 780000,
            "net_income": 250000
        },
        "cash_flow_statement": {
            "operating_activities": {
                "net_income": 250000,
                "adjustments": [
                    {"item": "Depreciation", "amount": 50000},
                    {"item": "Changes in Working Capital", "amount": -25000}
                ],
                "net_cash_from_operations": 275000
            },
            "investing_activities": {
                "items": [
                    {"item": "Purchase of Equipment", "amount": -100000}
                ],
                "net_cash_from_investing": -100000
            },
            "financing_activities": {
                "items": [
                    {"item": "Dividends Paid", "amount": -50000}
                ],
                "net_cash_from_financing": -50000
            },
            "net_change_in_cash": 125000,
            "cash_beginning": 25000,
            "cash_ending": 150000
        },
        "generated_at": datetime.utcnow().isoformat()
    }

@router.get("/")
def get_financial_statements(company_name: str = Query(...)):
    """Get list of generated financial statements"""
    try:
        # This would typically list saved financial statements
        # For now, return a sample list
        return {
            "statements": [
                {
                    "id": 1,
                    "period": "Q4",
                    "year": "2024",
                    "statement_type": "consolidated",
                    "generated_date": "2024-12-01T10:00:00",
                    "status": "final"
                },
                {
                    "id": 2,
                    "period": "Q3",
                    "year": "2024",
                    "statement_type": "consolidated",
                    "generated_date": "2024-09-01T10:00:00",
                    "status": "final"
                }
            ]
        }
        
    except Exception as e:
        print(f"Error getting financial statements: {e}")
        return {"statements": []}

@router.get("/export/{period}/{year}")
def export_financial_statements(
    period: str, 
    year: str, 
    format: str = Query("excel"), 
    company_name: str = Query(...)
):
    """Export financial statements to Excel or PDF"""
    try:
        if format not in ["excel", "pdf"]:
            raise HTTPException(status_code=400, detail="Format must be 'excel' or 'pdf'")
        
        # Generate sample export file
        if format == "excel":
            # Create temporary Excel file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                # Create sample Excel file with pandas
                df_bs = pd.DataFrame([
                    {"Account": "Cash", "Amount": 150000},
                    {"Account": "Accounts Receivable", "Amount": 85000},
                    {"Account": "Total Assets", "Amount": 235000}
                ])
                
                with pd.ExcelWriter(tmp_file.name, engine='openpyxl') as writer:
                    df_bs.to_excel(writer, sheet_name='Balance Sheet', index=False)
                
                return FileResponse(
                    path=tmp_file.name,
                    filename=f"financial_statements_{period}_{year}.xlsx",
                    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        
        else:  # PDF format
            raise HTTPException(status_code=501, detail="PDF export not implemented yet")
        
    except Exception as e:
        print(f"Error exporting financial statements: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export financial statements: {str(e)}"
        )

@router.get("/templates")
def get_statement_templates():
    """Get available financial statement templates"""
    return {
        "templates": [
            {
                "id": "ifrs_full",
                "name": "IFRS Full Financial Statements",
                "description": "Complete IFRS compliant financial statements",
                "components": ["balance_sheet", "income_statement", "cash_flow", "equity_changes", "notes"]
            },
            {
                "id": "ifrs_condensed",
                "name": "IFRS Condensed Statements",
                "description": "Condensed IFRS financial statements",
                "components": ["balance_sheet", "income_statement", "cash_flow"]
            },
            {
                "id": "management_reporting",
                "name": "Management Reporting Package",
                "description": "Internal management reporting format",
                "components": ["balance_sheet", "income_statement", "variance_analysis"]
            }
        ]
    }

@router.get("/{filename}/entity-breakdown")
def get_entity_breakdown(filename: str, company_name: str = Query(...)):
    """Get entity-level breakdown for consolidated statements"""
    try:
        # This would typically read from saved statement files
        # For now, return sample entity breakdown
        return {
            "filename": filename,
            "entity_breakdown": [
                {
                    "entity_code": "ENT_001",
                    "entity_name": "Parent Company",
                    "total_assets": 800000,
                    "total_liabilities": 300000,
                    "total_equity": 500000,
                    "net_income": 150000
                },
                {
                    "entity_code": "ENT_002", 
                    "entity_name": "Subsidiary A",
                    "total_assets": 400000,
                    "total_liabilities": 150000,
                    "total_equity": 250000,
                    "net_income": 75000
                }
            ],
            "consolidation_adjustments": [
                {
                    "description": "Intercompany eliminations",
                    "amount": -50000
                },
                {
                    "description": "Goodwill impairment",
                    "amount": -25000
                }
            ]
        }
        
    except Exception as e:
        print(f"Error getting entity breakdown: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get entity breakdown: {str(e)}"
        )
